"""
Excel 저장 모듈.
- 기존 파일을 로드하고 (공고번호 + 출처사이트) 기준으로 중복을 제거한 뒤 신규 행만 추가한다.
- 마감일이 7일 이내인 행은 주황색으로 강조한다.
- 공고링크와 첨부파일경로는 클릭 가능한 하이퍼링크로 삽입한다.
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Excel에 저장할 컬럼 순서 (내부 전달 키 _attachment_urls는 제외)
COLUMNS = [
    "공고번호",
    "공고명",
    "발주기관",
    "출처사이트",
    "공고일",
    "마감일시",
    "예산금액",
    "내용요약",
    "첨부파일수",
    "첨부파일경로",
    "공고링크",
    "수집일시",
]

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_URGENT_FILL = PatternFill("solid", fgColor="FF9900")   # 마감 7일 이내
_WRAP = Alignment(wrap_text=True, vertical="top")

# 열별 최적 너비 힌트 (문자 수 기준)
_COL_WIDTHS = {
    "공고번호":   18,
    "공고명":     45,
    "발주기관":   22,
    "출처사이트": 12,
    "공고일":     12,
    "마감일시":   18,
    "예산금액":   16,
    "내용요약":   50,
    "첨부파일수": 10,
    "첨부파일경로": 40,
    "공고링크":   40,
    "수집일시":   18,
}


def save_announcements(
    announcements: list[dict],
    config: dict,
    sheet_name: str = "공고목록",
) -> int:
    """
    공고 목록을 Excel 파일에 저장한다.

    Args:
        announcements: 저장할 공고 딕셔너리 리스트.
        config: 전체 config 딕셔너리.
        sheet_name: 저장할 시트 이름. 없으면 새로 생성한다.

    Returns:
        실제로 추가된 신규 행 수.
    """
    out_cfg = config.get("output", {})
    out_dir = Path(out_cfg.get("directory", "./output"))
    out_dir.mkdir(parents=True, exist_ok=True)

    if out_cfg.get("rolling_file", True):
        filepath = out_dir / out_cfg.get("filename", "announcements.xlsx")
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filepath = out_dir / f"announcements_{date_str}.xlsx"

    # 기존 파일 로드 or 새 워크북 생성
    if filepath.exists():
        wb = openpyxl.load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            existing_keys = _load_existing_keys(ws)
        else:
            ws = wb.create_sheet(sheet_name)
            _write_header(ws)
            existing_keys: set[tuple] = set()
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        _write_header(ws)
        existing_keys: set[tuple] = set()

    # 신규 공고만 필터링
    new_rows = [
        ann for ann in announcements
        if (ann.get("공고번호", ""), ann.get("출처사이트", "")) not in existing_keys
    ]

    if not new_rows:
        logger.info("신규 공고 없음, Excel 파일 변경 없음")
        return 0

    for ann in new_rows:
        _append_row(ws, ann)

    _apply_formatting(ws)
    wb.save(filepath)
    logger.info("Excel 저장 완료: %s (신규 %d건)", filepath, len(new_rows))
    return len(new_rows)


def _load_existing_keys(ws) -> set[tuple]:
    """기존 시트에서 (공고번호, 출처사이트) 조합을 읽어 중복 확인용 집합으로 반환."""
    keys: set[tuple] = set()
    header = [cell.value for cell in ws[1]]
    try:
        idx_num = header.index("공고번호")
        idx_src = header.index("출처사이트")
    except ValueError:
        return keys
    for row in ws.iter_rows(min_row=2, values_only=True):
        num = row[idx_num] if idx_num < len(row) else None
        src = row[idx_src] if idx_src < len(row) else None
        if num and src:
            keys.add((str(num), str(src)))
    return keys


def _write_header(ws):
    ws.append(COLUMNS)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    ws.row_dimensions[1].height = 20


def _append_row(ws, ann: dict):
    row_num = ws.max_row + 1
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        value = ann.get(col_name, "")

        if col_name == "공고링크" and value:
            cell.hyperlink = value
            cell.value = value
            cell.font = Font(color="0563C1", underline="single")
        elif col_name == "첨부파일경로" and value:
            # 로컬 폴더 경로를 file:/// 링크로 변환
            folder_uri = Path(value).as_uri()
            cell.hyperlink = folder_uri
            cell.value = value
            cell.font = Font(color="0563C1", underline="single")
        else:
            cell.value = value

        cell.alignment = _WRAP


def _apply_formatting(ws):
    """전체 시트에 열 너비 적용 및 긴급 행 강조."""
    # 열 너비
    for col_idx, col_name in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _COL_WIDTHS.get(col_name, 15)

    # 마감일 기준 긴급 강조 (헤더 제외)
    today = datetime.now().date()
    deadline_col_idx = COLUMNS.index("마감일시") + 1

    for row in ws.iter_rows(min_row=2):
        deadline_cell = row[deadline_col_idx - 1]
        deadline_str = str(deadline_cell.value or "")
        if _is_urgent(deadline_str, today):
            for cell in row:
                # 하이퍼링크 셀의 폰트는 건드리지 않고 배경색만 변경
                if cell.hyperlink is None:
                    cell.fill = _URGENT_FILL


def _is_urgent(deadline_str: str, today) -> bool:
    """마감일시 문자열이 오늘로부터 7일 이내이면 True."""
    if not deadline_str or deadline_str.strip() == "":
        return False
    # 다양한 날짜 형식 파싱 시도
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            deadline_date = datetime.strptime(deadline_str[:len(fmt)], fmt).date()
            return today <= deadline_date <= today + timedelta(days=7)
        except ValueError:
            continue
    return False
